"""
Legacy .xls → .xlsx conversion.

Two strategies, in order of preference:

  1. **COM** (Windows only) — drive the locally-installed Excel or WPS via
     `win32com.client`. Preserves formulas, images, conditional formatting,
     everything. Requires Microsoft Excel or WPS Office to be installed
     (most H3C 配置器 users have one or the other).

  2. **xlrd fallback** (pure Python, cross-platform) — read cell VALUES via
     `xlrd`, write a fresh .xlsx via openpyxl. Loses formulas, embedded
     images, and most formatting. We use it only as a last resort and warn
     loudly so the user understands the trade-off.

Callers should treat this as a one-shot: call `convert_xls_to_xlsx(src)`
and use the returned `.xlsx` path. Temp directories are cleaned up by
`format_quote` after it writes the final output.
"""

from __future__ import annotations

import logging
import platform
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class ConversionResult:
    src: Path
    dst: Path
    method: str             # "com" | "xlrd"
    temp_dir: Path | None   # set when dst lives in a temp dir we should rm later
    warnings: list[str]


def convert_xls_to_xlsx(
    src: Path | str,
    dst: Path | str | None = None,
    *,
    force_method: str | None = None,
) -> ConversionResult:
    """
    Convert a .xls file to .xlsx.

    Args:
        src: input .xls
        dst: output .xlsx; if None, written into a fresh temp dir (caller
             is responsible for cleaning that up via `result.temp_dir`)
        force_method: "com" / "xlrd" to bypass the auto-pick

    Returns a ConversionResult so callers know which method was used (and
    whether to warn the user about formula loss).
    """
    src = Path(src)
    if not src.exists():
        raise FileNotFoundError(src)
    if src.suffix.lower() != ".xls":
        raise ValueError(f"Expected .xls input, got {src.suffix!r}")

    temp_dir: Path | None = None
    if dst is None:
        temp_dir = Path(tempfile.mkdtemp(prefix="unis_xls_"))
        dst = temp_dir / (src.stem + ".xlsx")
    else:
        dst = Path(dst)
    dst.parent.mkdir(parents=True, exist_ok=True)

    method = force_method or _pick_method()
    warnings: list[str] = []

    if method == "com":
        try:
            _convert_via_com(src, dst)
            return ConversionResult(src, dst, "com", temp_dir, warnings)
        except Exception as exc:                                  # noqa: BLE001
            logger.warning("COM conversion failed (%s); falling back to xlrd", exc)
            warnings.append(
                f"Excel/WPS COM 调用失败,自动回退到 xlrd ({type(exc).__name__})。"
                f"原始公式可能丢失。"
            )
            method = "xlrd"

    if method == "xlrd":
        _convert_via_xlrd(src, dst)
        warnings.append(
            "用了 xlrd 纯 Python 回退路径 — 原 .xls 里的公式和图片不会保留,"
            "只保留单元格的数值/文本。如需保留公式,请装 Excel 或 WPS Office。"
        )
        return ConversionResult(src, dst, "xlrd", temp_dir, warnings)

    raise ValueError(f"Unknown conversion method: {method!r}")


def cleanup(result: ConversionResult) -> None:
    """Remove the temp dir produced by `convert_xls_to_xlsx` if any."""
    if result.temp_dir and result.temp_dir.exists():
        shutil.rmtree(result.temp_dir, ignore_errors=True)


# ---------------------------------------------------------------------------
# Strategy selection
# ---------------------------------------------------------------------------
def _pick_method() -> str:
    """COM on Windows when win32com is importable, otherwise xlrd."""
    if platform.system() != "Windows":
        return "xlrd"
    try:
        import win32com.client  # noqa: F401
        return "com"
    except ImportError:
        return "xlrd"


# ---------------------------------------------------------------------------
# COM path (Excel / WPS)
# ---------------------------------------------------------------------------
# Excel FileFormat constant for OpenXML (.xlsx). See xlOpenXMLWorkbook in
# the Office VBA reference. WPS exposes the same constant for compatibility.
_XL_OPEN_XML_WORKBOOK = 51


def _convert_via_com(src: Path, dst: Path) -> None:
    """
    Drive Excel/WPS via COM. Robust and lossless.

    Notes:
      - `win32com.client.Dispatch` re-uses an already-running Excel
        instance if one is open; we tolerate that.
      - `DisplayAlerts=False` suppresses the "compatibility checker" pop-up.
      - We always Quit() in finally, otherwise leaked excel.exe processes
        pile up across runs.
    """
    import pythoncom        # noqa: F401 — initializes COM threading state
    import win32com.client as win32

    # AbsolutePath required — COM resolves relative paths against the
    # Office process's own working dir, which is unpredictable.
    src_abs = str(src.resolve())
    dst_abs = str(dst.resolve())

    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False

    wb = None
    try:
        wb = excel.Workbooks.Open(
            src_abs,
            UpdateLinks=0,                  # don't try to update external refs
            ReadOnly=True,                  # belt-and-suspenders: never write back to .xls
            IgnoreReadOnlyRecommended=True,
        )
        wb.SaveAs(dst_abs, FileFormat=_XL_OPEN_XML_WORKBOOK)
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:                                         # noqa: BLE001
            pass
        try:
            excel.Quit()
        except Exception:                                         # noqa: BLE001
            pass


# ---------------------------------------------------------------------------
# xlrd fallback
# ---------------------------------------------------------------------------
def _convert_via_xlrd(src: Path, dst: Path) -> None:
    """
    Best-effort pure-Python fallback. Reads cell VALUES only — formulas
    are read as the cached display value, images are dropped entirely.
    """
    import xlrd
    from openpyxl import Workbook

    book = xlrd.open_workbook(str(src), formatting_info=False)
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name in book.sheet_names():
        src_sheet = book.sheet_by_name(sheet_name)
        dst_sheet = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name cap
        for r in range(src_sheet.nrows):
            for c in range(src_sheet.ncols):
                v = src_sheet.cell_value(r, c)
                if v == "":
                    continue
                # xlrd may give us numbers as float — preserve int-ness
                if isinstance(v, float) and v.is_integer():
                    v = int(v)
                dst_sheet.cell(row=r + 1, column=c + 1, value=v)
    wb.save(str(dst))


__all__ = ["convert_xls_to_xlsx", "cleanup", "ConversionResult"]
