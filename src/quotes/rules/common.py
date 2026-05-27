"""
Universal rules applied to EVERY quote, regardless of product category.

Rules in this module:
  - DropFixedColumns:  remove 产品名称 / 详细描述 / 要求提前报备周期 / 订单准备周期
                       columns from 价格汇总表
  - FillEmptyModel:    when the 产品型号 cell is empty, copy the first
                       UNIS model code out of the 描述 column
  - RemoveH3CLogo:     delete the H3C logo image in the top-left of
                       价格汇总表
"""

from __future__ import annotations

import re
from typing import TYPE_CHECKING

from ..inspector import find_header, iter_data_rows
from ..reader import has_sheet, require_sheet
from .base import QuoteRule, RuleResult

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook


# ---------------------------------------------------------------------------
# Rule 1 — Drop the 4 fixed columns the user always removes from 价格汇总表
# ---------------------------------------------------------------------------
class DropFixedColumns(QuoteRule):
    name = "drop_fixed_columns"
    description = "删除价格汇总表中固定的 4 列"

    # Header texts to remove. Exact match after `.strip()`.
    COLUMNS_TO_DROP = (
        "产品名称",
        "详细描述",
        "要求提前报备周期",
        "订单准备周期",
    )

    def apply(self, wb: "Workbook", context: dict) -> RuleResult:
        result = RuleResult(name=self.name)
        if not has_sheet(wb, "price_main"):
            result.warnings.append("没有 '价格汇总表' sheet,跳过")
            return result

        sheet = require_sheet(wb, "price_main")
        header = find_header(sheet)
        if header is None:
            result.warnings.append("没找到 '序号' 表头行,跳过")
            return result

        # Resolve target columns and delete in DESCENDING order so the
        # indexes of remaining columns stay valid as we go.
        targets: list[tuple[str, int]] = []
        for name in self.COLUMNS_TO_DROP:
            col = header.headers.get(name)
            if col is not None:
                targets.append((name, col))
        targets.sort(key=lambda t: -t[1])

        for name, col in targets:
            sheet.delete_cols(col, 1)
            result.changes.append(f"删除列 '{name}' (原列号 {col})")
        if not targets:
            result.warnings.append("4 个固定列一个都没找到 — 可能模板变了")
        else:
            result.applied = True
        return result


# ---------------------------------------------------------------------------
# Rule 2 — Fill empty 产品型号 from the 描述 column
# ---------------------------------------------------------------------------
class FillEmptyModel(QuoteRule):
    name = "fill_empty_model"
    description = "价格汇总表中 产品型号 为空时,从描述列提取 UNIS 型号填入"

    # First whitespace-separated token starting with UNIS/H3C, optionally
    # with a hyphen, then a model code. Matches all the patterns we see:
    #   "UNIS S5800-56T-EI-G 以太网交换机..."
    #   "UNIS Server R4930 G7 MAL0 CTO服务器..."
    #   "UNIS Storage XC20000F G6..."
    _MODEL_RE = re.compile(
        r"\b(UNIS(?:\s+(?:Server|Storage))?[\s\-]+[A-Z][A-Z0-9\-]*(?:\s+G\d)?)",
    )

    def apply(self, wb: "Workbook", context: dict) -> RuleResult:
        result = RuleResult(name=self.name)
        if not has_sheet(wb, "price_main"):
            result.warnings.append("没有 '价格汇总表' sheet,跳过")
            return result

        sheet = require_sheet(wb, "price_main")
        header = find_header(sheet)
        if header is None:
            result.warnings.append("没找到表头,跳过")
            return result

        model_col = header.headers.get("产品型号")
        desc_col = header.headers.get("描述")
        if model_col is None or desc_col is None:
            result.warnings.append(
                f"缺少 '产品型号' 或 '描述' 列(找到: {list(header.headers)})"
            )
            return result

        filled = 0
        for r in iter_data_rows(sheet, header):
            model_cell = sheet.cell(row=r, column=model_col)
            if model_cell.value not in (None, ""):
                continue
            desc_val = sheet.cell(row=r, column=desc_col).value
            if not desc_val:
                continue
            m = self._MODEL_RE.search(str(desc_val))
            if m:
                extracted = re.sub(r"\s+", " ", m.group(1).strip())
                model_cell.value = extracted
                row_marker = sheet.cell(row=r, column=1).value
                result.changes.append(f"R{r} ({row_marker}): 填入 '{extracted}'")
                filled += 1
            else:
                result.warnings.append(
                    f"R{r}: 描述里没找到 UNIS 型号 ({str(desc_val)[:40]}…)"
                )

        if filled:
            result.applied = True
        return result


# ---------------------------------------------------------------------------
# Rule 3 — Remove the H3C logo image
# ---------------------------------------------------------------------------
class RemoveH3CLogo(QuoteRule):
    name = "remove_h3c_logo"
    description = "删除左上角的 H3C logo 图片"

    def apply(self, wb: "Workbook", context: dict) -> RuleResult:
        result = RuleResult(name=self.name)
        if not has_sheet(wb, "price_main"):
            result.warnings.append("没有 '价格汇总表' sheet,跳过")
            return result
        sheet = require_sheet(wb, "price_main")

        # openpyxl stores images on sheet._images. We delete them all from
        # the 价格汇总表 sheet — the only image the configurator places
        # there is the H3C logo (no product photos on this sheet).
        before = len(sheet._images)
        if before == 0:
            result.warnings.append("没找到图片(可能已被手动删除)")
            return result
        sheet._images.clear()
        result.applied = True
        result.changes.append(f"删除了 {before} 张图片")
        return result


__all__ = ["DropFixedColumns", "FillEmptyModel", "RemoveH3CLogo"]
