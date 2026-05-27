"""
Locate things inside a quote workbook.

The H3C 配置器 leaves a lot of formatting cruft above the actual data
(empty rows, banner cells, multi-row titles), so we never use fixed row
numbers — we always search for the header row first, then operate
relative to it.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Iterator

from openpyxl.worksheet.worksheet import Worksheet


# Canonical header text expected to mark the data table. Each sheet has
# a different leading cell; we use the leftmost cell of the header row
# as a signal because it's always present and short ("序号").
HEADER_FIRST_CELL = "序号"


@dataclass
class HeaderInfo:
    row: int                           # 1-based Excel row
    headers: dict[str, int]            # header text → 1-based column index


def find_header(sheet: Worksheet, *, scan_rows: int = 60) -> HeaderInfo | None:
    """
    Walk the first `scan_rows` rows looking for the canonical header row.

    Returns None if not found — callers decide whether that's an error or
    a "this sheet has nothing for me" situation.
    """
    for r in range(1, scan_rows + 1):
        first = sheet.cell(row=r, column=1).value
        if first is None:
            continue
        if str(first).strip() != HEADER_FIRST_CELL:
            continue
        headers: dict[str, int] = {}
        for c in range(1, sheet.max_column + 1):
            cell_val = sheet.cell(row=r, column=c).value
            if cell_val is None:
                continue
            headers[str(cell_val).strip()] = c
        return HeaderInfo(row=r, headers=headers)
    return None


def iter_data_rows(sheet: Worksheet, header: HeaderInfo) -> Iterator[int]:
    """
    Yield 1-based row indices of DATA rows (everything below the header,
    skipping fully-blank rows, stopping at the "总计" summary row).
    """
    for r in range(header.row + 1, sheet.max_row + 1):
        # Stop at the "总计" row (typically in column 2 = 配置组名称)
        row_marker = sheet.cell(row=r, column=2).value
        if row_marker and str(row_marker).strip() == "总计":
            return
        # Skip fully blank rows but DON'T stop — H3C sometimes leaves
        # cosmetic blanks between groups.
        if _row_is_blank(sheet, r):
            continue
        yield r


def _row_is_blank(sheet: Worksheet, row: int) -> bool:
    for c in range(1, sheet.max_column + 1):
        v = sheet.cell(row=row, column=c).value
        if v not in (None, ""):
            return False
    return True


__all__ = ["HeaderInfo", "find_header", "iter_data_rows", "HEADER_FIRST_CELL"]
