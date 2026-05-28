"""
Load a quote workbook in a way that preserves formulas.

openpyxl opens .xlsx with formulas intact when `data_only=False`. We
never read computed values — when we need to display a price we read
either the formula string OR the cached value (whichever exists).

Legacy .xls (the older H3C 配置器 default) gets **auto-converted** to
.xlsx on the fly via `xls_convert.convert_xls_to_xlsx` — Excel/WPS COM
when available (lossless), xlrd fallback otherwise (values only).
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from .exceptions import SheetNotFoundError, UnsupportedFormatError
from .xls_convert import ConversionResult, convert_xls_to_xlsx

logger = logging.getLogger(__name__)

# H3C 配置器 export sheet name conventions — fixed across all quote files
# we've inspected. If H3C renames them in a future configurator release
# this is the single place to update.
SHEETS = {
    "cover":         "封面",
    "price_summary": "价格总表",
    "price_main":    "价格汇总表",      # main detail with 产品型号 / 描述 / 客户描述
    "price_detail":  "价格明细清单",    # line-item itemization, including BOM components
    "volume":        "体积重量功耗信息",
    "warranty":      "维保服务费明细清单",
    "notes":         "产品备注明细",
}


@dataclass
class LoadedQuote:
    """Workbook + provenance — needed when source was a converted .xls."""

    workbook: Workbook
    source_path: Path                       # the file the user actually gave us
    opened_path: Path                       # the file we actually opened (=source unless converted)
    conversion: ConversionResult | None     # set when .xls auto-conversion ran


def open_quote(
    path: str | Path,
    *,
    auto_convert_xls: bool = True,
) -> LoadedQuote:
    """
    Open a quote workbook for editing. Preserves formulas.

    If `path` is .xls and `auto_convert_xls=True` (default), we transparently
    convert to .xlsx via Excel/WPS COM (preferred, lossless) or xlrd fallback
    (values only). The returned LoadedQuote carries the ConversionResult so
    callers can warn users about formula loss and clean up the temp file.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(p)
    suffix = p.suffix.lower()

    conversion: ConversionResult | None = None
    if suffix == ".xls":
        if not auto_convert_xls:
            raise UnsupportedFormatError(
                f"{p.name}: 旧版 .xls 格式无法保留公式。请先用 Excel/WPS "
                f"另存为 .xlsx,或允许自动转换(auto_convert_xls=True)。"
            )
        logger.info("Auto-converting %s -> .xlsx", p.name)
        conversion = convert_xls_to_xlsx(p)
        opened_path = conversion.dst
        logger.info("Conversion done via %s -> %s", conversion.method, opened_path)
    elif suffix in (".xlsx", ".xlsm"):
        opened_path = p
    else:
        raise UnsupportedFormatError(
            f"{p.name}: 不支持的格式 {suffix!r}。期望 .xls / .xlsx / .xlsm。"
        )

    wb = load_workbook(
        filename=str(opened_path),
        data_only=False,     # keep formulas as formulas
        keep_vba=(opened_path.suffix.lower() == ".xlsm"),
    )
    return LoadedQuote(workbook=wb, source_path=p, opened_path=opened_path,
                       conversion=conversion)


def require_sheet(wb: Workbook, sheet_key: str):
    """Fetch a sheet by its logical key (see `SHEETS` above)."""
    if sheet_key not in SHEETS:
        raise KeyError(f"Unknown sheet key: {sheet_key!r}")
    name = SHEETS[sheet_key]
    if name not in wb.sheetnames:
        raise SheetNotFoundError(
            f"Workbook is missing sheet {name!r} (expected for {sheet_key})."
        )
    return wb[name]


def has_sheet(wb: Workbook, sheet_key: str) -> bool:
    return SHEETS.get(sheet_key) in wb.sheetnames


__all__ = ["open_quote", "require_sheet", "has_sheet", "SHEETS"]
