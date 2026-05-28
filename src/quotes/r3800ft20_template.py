"""
R3800FT20 G3 配置模板 loader.

The H3C 配置器 exports R3800FT20 G3 server quotes as a single bundled
CTO row in 价格明细清单 (one big SKU, no breakdown). The user separately
maintains a "配置模板" Excel file that lists every CTO component on its
own row, with a 数量 column they manually fill in to mark the ones they
selected.

This module reads that template, filters down to rows with 数量 > 0,
and returns them as TemplateItem records. The format rule
`fill_r3800ft20_template` then writes those items into the main quote.

File-discovery rules (same pattern as bom_lookup):
  - `quotes.r3800ft20_template_path` in config — supports glob
  - Files named with "已选型"/"已选" preferred over the bare template
    (the latter is the un-edited skeleton)
  - Among preferred candidates, most recent mtime wins
"""

from __future__ import annotations

import glob
import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# Rows whose 产品编码 OR 描述 cell matches one of these are footer / total
# markers in the template, not actual product items — skip them even if
# their 数量 column happens to be set.
_FOOTER_MARKERS = frozenset({"单台", "数量", "小计", "配置组小计", "总计"})


@dataclass(frozen=True)
class TemplateItem:
    code: str                  # 产品编码
    model: str                 # 产品型号
    product_code: str          # 产品代码
    description: str           # 描述
    qty: float                 # 数量
    list_price: float | None   # 目录单价 (may be missing in some templates)
    discount: float | None     # 折扣 (e.g. 1.0 = 100%; may be missing)


# ---------------------------------------------------------------------------
def load_template(path: Path) -> list[TemplateItem]:
    """
    Return all rows from the template's 价格明细清单 sheet where 数量 > 0,
    skipping footer / summary rows.
    """
    wb = load_workbook(str(path), data_only=True, read_only=True)
    try:
        if "价格明细清单" not in wb.sheetnames:
            raise ValueError(
                f"{path.name}: missing '价格明细清单' sheet "
                f"(available: {wb.sheetnames})"
            )
        ws = wb["价格明细清单"]

        # Locate the canonical header row
        header_r = None
        for r in range(1, 25):
            v = ws.cell(r, 1).value
            if v and str(v).strip() == "序号":
                header_r = r
                break
        if header_r is None:
            raise ValueError(f"{path.name}: no '序号' header row found")

        headers = {
            ws.cell(header_r, c).value: c
            for c in range(1, ws.max_column + 1)
            if ws.cell(header_r, c).value
        }
        qty_col = headers.get("数量")
        if qty_col is None:
            raise ValueError(f"{path.name}: no '数量' column")
        code_col = headers.get("产品编码")
        model_col = headers.get("产品型号")
        pcode_col = headers.get("产品代码")
        desc_col = headers.get("描述")
        price_col = headers.get("目录单价(RMB)") or headers.get("目录单价")
        discount_col = headers.get("折扣")

        items: list[TemplateItem] = []
        for r in range(header_r + 1, ws.max_row + 1):
            qty = ws.cell(r, qty_col).value
            if not isinstance(qty, (int, float)) or qty <= 0:
                continue

            code = _cell_str(ws, r, code_col)
            desc = _cell_str(ws, r, desc_col)

            # Skip footer / summary rows even if they have a quantity
            if code in _FOOTER_MARKERS or desc in _FOOTER_MARKERS:
                continue

            list_price: float | None = None
            if price_col:
                v = ws.cell(r, price_col).value
                if isinstance(v, (int, float)) and v > 0:
                    list_price = float(v)

            discount: float | None = None
            if discount_col:
                v = ws.cell(r, discount_col).value
                if isinstance(v, (int, float)) and v > 0:
                    discount = float(v)

            items.append(TemplateItem(
                code=code,
                model=_cell_str(ws, r, model_col),
                product_code=_cell_str(ws, r, pcode_col),
                description=desc,
                qty=float(qty),
                list_price=list_price,
                discount=discount,
            ))
    finally:
        wb.close()

    logger.info("Loaded %d items from R3800FT20 template %s",
                len(items), path.name)
    return items


def _cell_str(ws, r: int, c: int | None) -> str:
    if c is None:
        return ""
    v = ws.cell(r, c).value
    return str(v).strip() if v is not None else ""


# ---------------------------------------------------------------------------
def resolve_template_path(configured: Path | str | None) -> Path | None:
    """
    Resolve the configured template path. Supports glob; among matches,
    prefer files marked "已选型" / "已选" (the user-curated copy) over
    the bare template skeleton.
    """
    if not configured:
        return None
    pattern = str(configured)
    if not any(ch in pattern for ch in "*?["):
        p = Path(pattern)
        return p if p.exists() else None

    matches = [Path(p) for p in glob.glob(pattern) if Path(p).is_file()]
    if not matches:
        return None
    selected = [p for p in matches if "已选" in p.name]
    pool = selected or matches
    return max(pool, key=lambda p: p.stat().st_mtime)


__all__ = ["TemplateItem", "load_template", "resolve_template_path"]
