"""
IT 产品 BOM lookup (IT产品BOM编码YYYYMMDD.xlsx).

The H3C 配置器 ships server quotes with a default OEM service line:
    "OEM服务器3年5×9下一工作日现场支持(含硬盘不返还) -附专属服务"

Per the user's working spec, this line must be swapped for the proper
"3年7×24×NBD维保(含硬盘介质保留)" entry from the IT产品BOM workbook.
This module is the lookup half — it loads the workbook, indexes the
OEM IT产品 sheet by description, and serves results to the rule that
mutates the quote.

The BOM file's path is taken from `config.quotes.bom_path` (defaults to
a glob in the user's Downloads folder so they can drop new monthly
versions without editing config).
"""

from __future__ import annotations

import glob
import logging
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Column layout of the "OEM IT产品" sheet (1-based, matches Excel A-H).
_COL_PRODUCT_LINE = 1   # 产品线         e.g. "HC"
_COL_MODEL        = 2   # 产品型号       e.g. "UNIS Server R4930 G7"
_COL_SERVICE_BOM  = 3   # 服务BOM        e.g. "8813L0H3"          ← 产品编码
_COL_EXT_MODEL    = 4   # 对外型号       e.g. "SV-MA-BS8-WD-..."   ← 产品代码
_COL_DESC         = 5   # 对外中文描述   ← the key we lookup on
_COL_LIST_PRICE   = 6   # 目录价格

_SHEET_NAME = "OEM IT产品"


@dataclass(frozen=True)
class BomEntry:
    code: str          # 产品编码 (服务BOM)
    model: str         # 产品型号
    ext_code: str      # 产品代码 (对外型号)
    description: str   # 对外中文描述
    list_price: float


# ---------------------------------------------------------------------------
def load_bom(path: Path) -> list[BomEntry]:
    """Load every row of the OEM IT产品 sheet into a list of BomEntry."""
    wb = load_workbook(str(path), data_only=True, read_only=True)
    if _SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"BOM workbook {path.name!r} is missing sheet {_SHEET_NAME!r}. "
            f"Available: {wb.sheetnames}"
        )
    ws = wb[_SHEET_NAME]
    entries: list[BomEntry] = []
    for r in range(2, ws.max_row + 1):
        desc = ws.cell(r, _COL_DESC).value
        if not desc:
            continue
        try:
            price = float(ws.cell(r, _COL_LIST_PRICE).value or 0)
        except (TypeError, ValueError):
            price = 0.0
        entries.append(BomEntry(
            code=str(ws.cell(r, _COL_SERVICE_BOM).value or "").strip(),
            model=str(ws.cell(r, _COL_MODEL).value or "").strip(),
            ext_code=str(ws.cell(r, _COL_EXT_MODEL).value or "").strip(),
            description=str(desc).strip(),
            list_price=price,
        ))
    wb.close()
    logger.info("Loaded %d BOM entries from %s", len(entries), path.name)
    return entries


# ---------------------------------------------------------------------------
@lru_cache(maxsize=8)
def _load_cached(path_str: str, mtime: float) -> tuple[BomEntry, ...]:
    """LRU-cached load keyed on (path, mtime) so reloads pick up file changes."""
    return tuple(load_bom(Path(path_str)))


def get_bom(path: Path) -> list[BomEntry]:
    """Cached load — returns a fresh list on each call but skips disk I/O."""
    return list(_load_cached(str(path.resolve()), path.stat().st_mtime))


# ---------------------------------------------------------------------------
def find_warranty(
    entries: list[BomEntry],
    server_model: str,
    *,
    years: int = 3,
    hdd_retained: bool = True,
) -> BomEntry | None:
    """
    Find the "<server_model> Y年7×24×NBD维保[(含硬盘介质保留)]" row.

    Defaults match the user's Q3 spec: 3 年, 含硬盘介质保留.
    """
    base = f"{server_model} {years}年7×24×NBD维保"
    expected = base + ("(含硬盘介质保留)" if hdd_retained else "")

    # Exact match first.
    for e in entries:
        if e.description == expected:
            return e

    # Forgiving fallback: starts-with on the base, with the right HDD flag.
    for e in entries:
        d = e.description
        if not d.startswith(base):
            continue
        is_hdd = "含硬盘介质保留" in d
        if is_hdd == hdd_retained:
            return e
    return None


# ---------------------------------------------------------------------------
def resolve_bom_path(configured: Path | str | None) -> Path | None:
    """
    Resolve the configured BOM path, supporting glob patterns so the user
    can write `IT产品BOM编码*.xlsx` and get the most recent file.

    Returns None if nothing matches — the caller should warn and skip
    the swap rule.
    """
    if not configured:
        return None
    pattern = str(configured)
    if any(ch in pattern for ch in "*?["):
        matches = [Path(p) for p in glob.glob(pattern)]
        matches = [p for p in matches if p.is_file()]
        if not matches:
            return None
        # Most recent mtime wins
        return max(matches, key=lambda p: p.stat().st_mtime)
    p = Path(pattern)
    return p if p.exists() else None


__all__ = [
    "BomEntry",
    "load_bom",
    "get_bom",
    "find_warranty",
    "resolve_bom_path",
]
