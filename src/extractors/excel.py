"""Excel .xlsx extractor — every sheet becomes one table; meta tracks sheet count."""

from __future__ import annotations

from pathlib import Path

from .base import ExtractedContent, ExtractedTable


def extract_excel(path: Path) -> ExtractedContent:
    """
    Extract every worksheet as a table.

    We deliberately keep the layout flat — no attempt to merge cells or detect
    "the real header row". Downstream parsers can apply heuristics if needed.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    tables: list[ExtractedTable] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v).strip() for v in row]
            if any(cells):                        # skip fully-blank rows
                rows.append(cells)
        if len(rows) >= 2:
            tables.append(ExtractedTable(rows=rows, sheet=sheet_name))

    # Provide a text fallback so downstream prompts always have *something*.
    text = "\n\n".join(
        f"# Sheet: {t.sheet}\n" + "\n".join(" | ".join(r) for r in t.rows)
        for t in tables
    )

    return ExtractedContent(
        source=path,
        kind="xlsx",
        text=text,
        tables=tables,
        meta={"sheet_count": len(tables)},
    )


__all__ = ["extract_excel"]
